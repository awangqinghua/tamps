#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2022/9/26 17:51
# @Author   : qinghua
# @Email    : 867075698@qq.com
# @File     : do_excels.py
# @Software : PyCharm

from openpyxl import load_workbook


class DeExcels:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row):
            test_datas = {}
            test_datas["video_path"] = sheet.cell(i+1,8).value
            test_data.append(test_datas)
        return test_data


test = DeExcels(r"D:\tamps\data\大型车负.xlsx", "Sheet1").get_data()



if __name__ == '__main__':

    pass
    print(test)