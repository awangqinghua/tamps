#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2022/9/26 16:44
# @Author   : qinghua
# @Email    : 867075698@qq.com
# @File     : mysql_s.py
# @Software : PyCharm


from openpyxl import load_workbook


class DeExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row):
            test_datas = {}
            test_datas["ids"] = sheet.cell(i + 1, 1).value
            test_datas["video_path"] = sheet.cell(i+1,2).value
            test_data.append(test_datas)
        return test_data


test = DeExcel(r"D:\tamps\data\大型车占用主车道.xlsx", "Sheet1").get_data()


if __name__ == '__main__':
    pass

