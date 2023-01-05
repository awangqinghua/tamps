#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time :2022/3/14 17:18
# @Author :wangqinghua
# @File : do_excel.py
# @Software : PyCharm


from openpyxl import load_workbook


class DeExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row):
            sub_data={}
            sub_data["布控id"] = sheet.cell(i+1,3).value
            test_data.append(sub_data)
        return test_data


test = DeExcel(r"D:\tamps\data\data1130\1206.xlsx", "1206").get_data()


def test_key():
    res = []
    for i in test:
        res.append(i["布控id"])
    return res


if __name__ == '__main__':
    print(test_key())
    print(len(test_key()))


