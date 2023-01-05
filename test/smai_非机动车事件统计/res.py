#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2022/12/16 9:49
# @Author   : wqh
# @Email    : 867075698@qq.com
# @File     : res.py
# @Software : PyCharm


from openpyxl import load_workbook

test = load_workbook("../../data/data1214/布控任务.xlsx")
sheet = test['布控任务']

print(type(sheet.cell(2, 13).value))
# print(sheet.max_row)
# print(sheet.max_column)

# 从excel获取到的数据是什么类型
# 从excel读取数据：数字还是数字  其他都是字符串类型
# print("method:{}，类型是{}".format(sheet.cell(1,1).value,type(sheet.cell(1,1).value)))
# print("url:{}，类型是{}".format(sheet.cell(1,2).value,type(sheet.cell(1,2).value)))
# print("data:{}，类型是{}".format(sheet.cell(1,3).value,type(sheet.cell(1,3).value)))
# print("code:{}，类型是{}".format(sheet.cell(1,4).value,type(sheet.cell(1,4).value)))
