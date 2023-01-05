#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time :2022/3/14 17:18
# @Author :qh
# @File : do_excel.py
# @Software : PyCharm


from openpyxl import load_workbook


class DeExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name, read_only=True)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row):  # 一直变化的是行，所以要获取最大行
            test_datas = {"任务编号": sheet.cell(i + 1, 1).value}
            test_data.append(test_datas)
        return test_data


test = DeExcel("../../data/data1219/布控任务.xlsx", "布控任务").get_data()


def test_key():
    res = []
    for i in test:
        res.append(i["任务编号"])
    return res

test_key()


if __name__ == '__main__':
    pass
    # 获取数据
    print(test_key())
    print(len(test_key()))






