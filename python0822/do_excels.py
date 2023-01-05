#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time :2022/3/14 17:18
# @Author :wangqinghua
# @File : do_excel.py
# @Software : PyCharm

from openpyxl import load_workbook


class DeExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []
        for i in range(1, sheet.max_row):
            sub_data={}
            sub_data["任务编号"] = sheet.cell(i+1, 1).value
            test_data.append(sub_data)
        return test_data


def test_key():
    res = []
    for i in test:
        res.append(i["任务编号"])
    return res


test = DeExcel("../data/datas/布控任务1.xlsx", "布控任务1").get_data()
# print(te.txt)
print(len(test))



